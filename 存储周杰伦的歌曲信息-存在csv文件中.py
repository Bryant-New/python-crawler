#  coding: gbk
# Python爬虫过程-获取数据-解析数据-提取数据-存储数据
import requests
import openpyxl

# csv文件写入：1.创建文件-调用open函数2.创建对象-借助writer函数3.写入内容-writerow方法4.关闭文件
# import csv
# csv_file = open('demo.csv','w',newline='')
# writer = csv.writer(csv_file)
# writer.writerow(['电影','豆瓣评分'])
# csv_file.close()

# csv文件读取：1.打开文件-调用open函数2.创建对象-借助reader对象3.读取内容-遍历reader对象4.打印内容
# import csv
# csv_file = open('demo.csv','r',newline='')
# reader=csv.reader(csv_file)
# for row in reader:
#     print(row)

# Excel文件写入：1.创建工作薄-openpyxl.Workbook()2.获取工作表-workbook中的active属性3.操作单元格-append4.保存工作薄-save
# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title ='new title'
# sheet['A1'] = '漫威宇宙'
# rows = [['美国队长','钢铁侠','蜘蛛侠','雷神'],['是','漫威','宇宙', '经典','人物']]
# for i in rows:
#     sheet.append(i)
# print(rows)
# wb.save('Marvel.xlsx')

# Excel文件读取：1.创建工作薄-openpyxl.Workbook()2.获取工作表-workbook对象的键3.读取单元格-借助单元格value属性4.打印单元格
# import openpyxl
# wb = openpyxl.load_workbook('Marvel.xlsx')
# sheet = wb['new title']
# sheetname = wb.sheetnames
# print(sheetname)
# A1_value = sheet['A1'].value
# print(A1_value)
