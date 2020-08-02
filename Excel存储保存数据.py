# 引入对Excel表格文件操作的库,openpyxl模块的官方文档：https://openpyxl.readthedocs.io/en/stable/
import openpyxl
# 利用openpyxl.Workbook()创建新的workbook对象，即新的空的Excel文件
wb = openpyxl.Workbook()
# 获取新建空的workbook的活动表
sheet = wb.active
# 对工作表-worksheet进行重命名
sheet.title = '测试'
# 操作单元格，往单元格里写入内容
# 往第一个工作表的A1单元格，写入漫威宇宙
sheet['A1'] = '漫威宇宙'
# 往表格中添加一行内容，使用append函数
row =['美国队长', '钢铁侠', '蜘蛛侠']
sheet.append(row)
# 一次性写入多行内容-先将多行内容写成列表，再放进大列表，赋值给rows
rows = [['美国队长', '钢铁侠', '蜘蛛侠'], ['是', '漫威', '宇宙', '经典', '人物']]
# 遍历rows.同时把遍历的内容添加到表格里-实现多行写入
for i in rows:
    sheet.append(i)
print(rows)
# 往excel文件中写入数据后，保存Excel文件,杨光-Excel的名字
wb.save('杨光.xlsx')
# 读取Excel里的数据
# 调用openpyxl.load_workbook()函数，打开“杨光.xlsx”文件。
wb = openpyxl.load_workbook('杨光.xlsx')
# 获取“杨光.xlsx”工作薄中名为“测试”的工作表。
sheet = wb['测试']
# 获取工作薄所有工作表的名字，并打印
sheetname = wb.sheetnames
print(sheetname)
# 把“测试”工作表中A1单元格赋值给A1_cell，再利用单元格value属性，就能打印出A1单元格的值。
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)